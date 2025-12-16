from openpyxl import load_workbook
import os

filepath = "data/TB_BARANG.xlsx"

if os.path.exists(filepath):
    wb = load_workbook(filepath)
    ws = wb.active
    
    # Print header
    print("HEADER ROW:")
    for col_idx, cell in enumerate(ws[1], start=1):
        print(f"  Col {col_idx}: {cell.value}")
    
    # Cari kolom KODE dan NAMA
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[cell.value.strip().upper()] = col_idx
    
    kode_col = headers.get('KODE_BARANG') or headers.get('KODE')
    nama_col = headers.get('NAMA_BARANG') or headers.get('NAMA')
    
    print(f"\nKODE column index: {kode_col}")
    print(f"NAMA column index: {nama_col}")
    
    # Print first 10 rows
    print(f"\nFirst 10 data rows:")
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=11, values_only=True), 1):
        if kode_col and nama_col:
            kode = row[kode_col - 1] if kode_col <= len(row) else None
            nama = row[nama_col - 1] if nama_col <= len(row) else None
            print(f"  Row {i}: KODE={kode} (type: {type(kode).__name__}), NAMA={nama}")
    
    # Count total kode
    total_kode = 0
    all_kode = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if kode_col:
            kode = row[kode_col - 1] if kode_col <= len(row) else None
            if kode:
                kode_str = str(int(kode)) if isinstance(kode, (int, float)) else str(kode).strip()
                total_kode += 1
                all_kode.append(kode_str)
    
    print(f"\nTotal kode dalam file: {total_kode}")
    print(f"Last 10 kode: {all_kode[-10:] if len(all_kode) > 10 else all_kode}")
    
    # Cek apakah kode 8992982206001 ada
    print(f"\nSearching for kode '8992982206001':")
    if '8992982206001' in all_kode:
        print(f"  FOUND!")
    else:
        print(f"  NOT FOUND in {total_kode} kode")
else:
    print(f"File not found: {filepath}")
