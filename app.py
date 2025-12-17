from flask import Flask, render_template, request, jsonify, send_file, session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
import os
from datetime import datetime
from urllib.parse import quote
import sqlite3
import uuid
import json

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.secret_key = 'stok-opname-secret-key-2025'  # Untuk session management

# Database setup
DB_FILE = "data/stok_opname.db"
DB_ENABLED = False

def init_db():
    """Initialize SQLite database"""
    global DB_ENABLED
    try:
        os.makedirs(os.path.dirname(DB_FILE) or '.', exist_ok=True)
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute('''CREATE TABLE IF NOT EXISTS so_sessions (
            id TEXT PRIMARY KEY,
            nama_area TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS so_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id TEXT,
            kode_barang TEXT,
            nama_barang TEXT,
            stok_real INTEGER,
            nama_area TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(session_id) REFERENCES so_sessions(id)
        )''')
        conn.commit()
        conn.close()
        DB_ENABLED = True
        print("✅ Database initialized successfully")
    except Exception as e:
        print(f"⚠️ Database initialization failed: {str(e)}")
        print("🔄 Falling back to in-memory storage (Vercel ephemeral mode)")
        DB_ENABLED = False

init_db()

# Fallback: in-memory storage untuk Vercel
in_memory_data = {}

def get_session_id():
    """Get or create session ID"""
    if 'so_session_id' not in session:
        session['so_session_id'] = str(uuid.uuid4())
    return session['so_session_id']

def get_so_data(session_id):
    """Get SO data dari database atau in-memory"""
    if DB_ENABLED:
        conn = sqlite3.connect(DB_FILE)
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        c.execute('SELECT * FROM so_items WHERE session_id = ? ORDER BY id', (session_id,))
        rows = c.fetchall()
        conn.close()
        return [dict(row) for row in rows]
    else:
        # Fallback to in-memory
        if session_id not in in_memory_data:
            in_memory_data[session_id] = []
        return in_memory_data[session_id]

def add_so_item(session_id, kode_barang, nama_barang, stok_real, nama_area):
    """Add item ke database atau in-memory"""
    if DB_ENABLED:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute('''INSERT INTO so_items (session_id, kode_barang, nama_barang, stok_real, nama_area)
                     VALUES (?, ?, ?, ?, ?)''',
                  (session_id, kode_barang, nama_barang, stok_real, nama_area))
        conn.commit()
        conn.close()
    else:
        # Fallback to in-memory
        if session_id not in in_memory_data:
            in_memory_data[session_id] = []
        item_id = len(in_memory_data[session_id]) + 1
        in_memory_data[session_id].append({
            'id': item_id,
            'kode_barang': kode_barang,
            'nama_barang': nama_barang,
            'stok_real': stok_real,
            'nama_area': nama_area
        })

def delete_so_item(session_id, item_id):
    """Delete item dari database atau in-memory"""
    if DB_ENABLED:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute('DELETE FROM so_items WHERE id = ? AND session_id = ?', (item_id, session_id))
        conn.commit()
        conn.close()
    else:
        # Fallback to in-memory
        if session_id in in_memory_data:
            in_memory_data[session_id] = [item for item in in_memory_data[session_id] if item['id'] != item_id]

def reset_so_data(session_id):
    """Reset semua data untuk session"""
    if DB_ENABLED:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute('DELETE FROM so_items WHERE session_id = ?', (session_id,))
        conn.commit()
        conn.close()
    else:
        # Fallback to in-memory
        if session_id in in_memory_data:
            in_memory_data[session_id] = []

# Folder untuk baca master barang (TB_BARANG.xlsx)
DATA_FOLDER = "data"
if not os.path.exists(DATA_FOLDER):
    os.makedirs(DATA_FOLDER)

# Folder untuk simpan hasil SO - gunakan /tmp untuk Vercel, uploads untuk local
import tempfile
try:
    # Coba gunakan uploads folder (untuk local development)
    UPLOADS_FOLDER = "uploads"
    if not os.path.exists(UPLOADS_FOLDER):
        os.makedirs(UPLOADS_FOLDER)
except (OSError, PermissionError):
    # Fallback ke /tmp untuk Vercel (read-only filesystem)
    UPLOADS_FOLDER = tempfile.gettempdir()

def load_master_barang():
    """Membaca master barang dari TB_BARANG.xlsx dengan atribut lengkap"""
    try:
        filepath = os.path.join(DATA_FOLDER, "TB_BARANG.xlsx")
        
        if not os.path.exists(filepath):
            return {}
        
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Cari index kolom dari header (case-insensitive)
        headers = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value:
                headers[cell.value.strip().upper()] = col_idx
        
        # Cek berbagai nama kolom yang mungkin (prioritas: KODE, NAMA)
        kode_col = headers.get('KODE') or headers.get('KODE_BARANG')
        nama_col = headers.get('NAMA') or headers.get('NAMA_BARANG')
        
        # Kolom opsional
        stok_col = headers.get('STOK')
        satuan_col = headers.get('SATUAN')
        kategori_col = headers.get('KATEGORI')
        
        if not kode_col or not nama_col:
            print("Error: Kolom KODE atau NAMA tidak ditemukan")
            return {}
        
        master = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            kode = row[kode_col - 1] if kode_col <= len(row) and row[kode_col - 1] else None
            nama = row[nama_col - 1] if nama_col <= len(row) and row[nama_col - 1] else None
            
            if kode and nama:
                # Normalize kode: convert ke string, strip whitespace
                kode = str(int(kode)) if isinstance(kode, (int, float)) else str(kode).strip()
                nama = str(nama).strip()
                
                # Ambil atribut tambahan
                stok = row[stok_col - 1] if stok_col and stok_col <= len(row) else None
                satuan = row[satuan_col - 1] if satuan_col and satuan_col <= len(row) else 'Unit'
                kategori = row[kategori_col - 1] if kategori_col and kategori_col <= len(row) else None
                
                master[kode] = {
                    'nama': nama,
                    'stok': stok if stok else 0,
                    'satuan': satuan if satuan else 'Unit',
                    'kategori': kategori
                }
        
        return master
    except Exception as e:
        print(f"Error membaca TB_BARANG.xlsx: {str(e)}")
        return {}

@app.route('/')
def index():
    """Halaman utama"""
    return render_template('index.html')

@app.route('/api/master-barang')
def api_master_barang():
    """API untuk mendapatkan semua master barang"""
    master = load_master_barang()
    print(f"DEBUG: Loaded {len(master)} items from master barang")
    if master:
        print(f"DEBUG: First 5 items: {list(master.items())[:5]}")
    return jsonify(master)

@app.route('/api/validasi-kode', methods=['POST'])
def api_validasi_kode():
    """API untuk validasi kode_barang"""
    data = request.get_json()
    kode = data.get('kode', '').strip()
    
    master = load_master_barang()
    
    if kode in master:
        item = master[kode]
        return jsonify({
            'valid': True,
            'nama_barang': item['nama'] if isinstance(item, dict) else item,
            'stok': item['stok'] if isinstance(item, dict) else None,
            'satuan': item['satuan'] if isinstance(item, dict) else 'Unit',
            'kategori': item['kategori'] if isinstance(item, dict) else None
        })
    else:
        return jsonify({
            'valid': False,
            'message': f'Kode {kode} tidak ditemukan di master barang'
        })

@app.route('/api/tambah-so', methods=['POST'])
def api_tambah_so():
    """API untuk menambah data SO - OPTIMIZED"""
    session_id = get_session_id()
    data = request.get_json()
    kode = data['kode_barang']
    stok_real = data['stok_real']
    nama_barang = data['nama_barang']
    nama_area = data['nama_area']
    
    # Quick duplicate check
    so_data = get_so_data(session_id)
    for item in so_data:
        if item['kode_barang'] == kode:
            return jsonify({'success': False, 'message': f'Duplikat kode {kode}'}), 400
    
    # Add to database
    add_so_item(session_id, kode, nama_barang, stok_real, nama_area)
    
    return jsonify({'success': True}), 200

@app.route('/api/daftar-so')
def api_daftar_so():
    """API untuk mendapatkan daftar SO"""
    session_id = get_session_id()
    so_data = get_so_data(session_id)
    return jsonify(so_data)

@app.route('/api/share-whatsapp', methods=['POST'])
def api_share_whatsapp():
    """API untuk generate WhatsApp share link"""
    session_id = get_session_id()
    so_data = get_so_data(session_id)
    
    if not so_data:
        return jsonify({'success': False, 'message': 'Tidak ada data untuk dibagikan'}), 400
    
    # Ambil nama area
    nama_area = so_data[0].get('nama_area', 'Area Tidak Ditentukan') if so_data else 'Area Tidak Ditentukan'
    
    # Buat pesan dengan format yang lebih rapi
    pesan = f"📦 *STOCK OPNAME - {nama_area}*\n"
    pesan += f"📅 Tanggal: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n"
    pesan += f"{'='*50}\n\n"
    
    pesan += "*📋 DETAIL ITEM:*\n"
    for idx, item in enumerate(so_data, 1):
        pesan += f"{idx}. {item['kode_barang']} - {item['nama_barang']}\n"
        pesan += f"   ▸ Stok: {item['stok_real']} unit\n"
    
    pesan += f"\n{'='*50}\n"
    pesan += f"📊 *RINGKASAN:*\n"
    pesan += f"✓ Total Item: {len(so_data)}\n"
    pesan += f"✓ Total Stok: {sum(item['stok_real'] for item in so_data)} unit\n\n"
    pesan += f"{'='*50}\n"
    pesan += f"📥 *UNTUK FILE EXCEL:*\n"
    pesan += f"1️⃣ Klik tombol 'Export & Kirim WA'\n"
    pesan += f"2️⃣ File otomatis download\n"
    pesan += f"3️⃣ Upload file ke chat WhatsApp"
    
    # Encode pesan untuk URL
    encoded_pesan = quote(pesan)
    
    # WhatsApp link (tanpa nomor, user bisa pilih contact)
    whatsapp_link = f"https://wa.me/?text={encoded_pesan}"
    
    return jsonify({
        'success': True,
        'link': whatsapp_link,
        'pesan': pesan
    })

@app.route('/api/hapus-so/<int:item_id>', methods=['DELETE'])
def api_hapus_so(item_id):
    """API untuk menghapus data SO"""
    session_id = get_session_id()
    
    try:
        delete_so_item(session_id, item_id)
        so_data = get_so_data(session_id)
        return jsonify({'success': True, 'data': so_data})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400

@app.route('/api/reset-so', methods=['POST'])
def api_reset_so():
    """API untuk reset semua data SO"""
    session_id = get_session_id()
    reset_so_data(session_id)
    return jsonify({'success': True, 'message': 'Data SO direset'})

def buat_excel_so_memory(session_id):
    """Helper function untuk membuat Excel SO - return sebagai BytesIO (tanpa simpan file)"""
    from io import BytesIO
    
    so_data = get_so_data(session_id)
    # Ambil nama area dari data pertama (semua sama area)
    nama_area = so_data[0].get('nama_area', 'Tidak Ditentukan') if so_data else 'Tidak Ditentukan'
    
    # Buat workbook baru
    wb = Workbook()
    ws = wb.active
    ws.title = "Hasil SO"
    
    # Info area
    ws['A1'] = "Nama Area:"
    ws['B1'] = nama_area
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    
    ws['A2'] = "Tanggal:"
    ws['B2'] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    ws['A2'].font = Font(bold=True)
    
    # Header
    headers = ["NO", "KODE_BARANG", "NAMA_BARANG", "STOK_REAL"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Data
    for row_idx, item in enumerate(so_data, start=5):
        ws.cell(row=row_idx, column=1).value = row_idx - 4  # No
        ws.cell(row=row_idx, column=2).value = item['kode_barang']
        ws.cell(row=row_idx, column=3).value = item['nama_barang']
        ws.cell(row=row_idx, column=4).value = item['stok_real']
    
    # Set column width
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    
    nama_area_clean = nama_area.replace(' ', '_').replace('/', '_')
    # Include timestamp
    timestamp = datetime.now().strftime('%d%m%Y_%H%M%S')
    filename = f"{nama_area_clean}_SO_{timestamp}.xlsx"
    
    # Return sebagai BytesIO (in-memory) tanpa simpan file
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output, filename, nama_area

@app.route('/api/export-excel', methods=['GET', 'POST'])
def api_export_excel():
    """API untuk export data SO ke Excel - support GET dan POST"""
    session_id = get_session_id()
    so_data = get_so_data(session_id)
    
    if not so_data:
        return jsonify({'success': False, 'message': 'Tidak ada data untuk diekspor'}), 400
    
    excel_io, filename, _ = buat_excel_so_memory(session_id)
    return send_file(excel_io, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/download-file/<filename>', methods=['GET'])
def api_download_file(filename):
    """API untuk download file dari uploads folder"""
    try:
        # Security: prevent directory traversal
        if '..' in filename or '/' in filename:
            return jsonify({'success': False, 'message': 'Invalid filename'}), 400
        
        filepath = os.path.join(UPLOADS_FOLDER, filename)
        
        # Double check path security
        if not os.path.abspath(filepath).startswith(os.path.abspath(UPLOADS_FOLDER)):
            return jsonify({'success': False, 'message': 'Invalid file path'}), 400
        
        if not os.path.exists(filepath):
            return jsonify({'success': False, 'message': 'File tidak ditemukan'}), 404
        
        return send_file(filepath, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        print(f"Download error: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 400

@app.route('/api/share-excel-whatsapp', methods=['POST'])
def api_share_excel_whatsapp():
    """API untuk share Excel ke WhatsApp dengan nomor tertentu"""
    session_id = get_session_id()
    so_data = get_so_data(session_id)
    
    if not so_data:
        return jsonify({'success': False, 'message': 'Tidak ada data untuk dibagikan'}), 400
    
    # Langsung download tanpa simpan file
    excel_io, filename, nama_area = buat_excel_so_memory(session_id)
    
    # Format nomor WhatsApp (remove semua karakter non-digit)
    nomor_wa = request.get_json().get('nomor', '+62 851-1731-0261')
    nomor_wa_clean = ''.join(filter(str.isdigit, nomor_wa))
    
    # Ganti 0 di awal dengan 62
    if nomor_wa_clean.startswith('0'):
        nomor_wa_clean = '62' + nomor_wa_clean[1:]
    elif not nomor_wa_clean.startswith('62'):
        nomor_wa_clean = '62' + nomor_wa_clean
    
    # Generate download link - direct download endpoint
    download_link = f"{request.host_url.rstrip('/')}/api/export-excel"
    
    # Pesan WhatsApp dengan detail item + link download
    pesan = f"📦 *STOCK OPNAME - {nama_area}*\n"
    pesan += f"📅 Tanggal: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n"
    pesan += f"{'='*50}\n\n"
    
    # Detail item
    pesan += "*📋 DETAIL ITEM:*\n"
    for idx, item in enumerate(so_data, 1):
        pesan += f"{idx}. {item['kode_barang']} - {item['nama_barang']}\n"
        pesan += f"   ▸ Stok: {item['stok_real']} unit\n"
    
    pesan += f"\n{'='*50}\n"
    pesan += f"📊 *RINGKASAN:*\n"
    pesan += f"✓ Total Item: {len(so_data)}\n"
    pesan += f"✓ Total Stok: {sum(item['stok_real'] for item in so_data)} unit\n\n"
    pesan += f"{'='*50}\n"
    pesan += f"📥 *DOWNLOAD FILE EXCEL:*\n"
    pesan += f"🔗 {download_link}\n\n"
    pesan += f"✅ Klik link di atas untuk download, lalu upload ke chat"
    
    encoded_pesan = quote(pesan)
    
    # WhatsApp link dengan nomor tertentu
    whatsapp_link = f"https://wa.me/{nomor_wa_clean}?text={encoded_pesan}"
    
    return jsonify({
        'success': True,
        'link': whatsapp_link,
        'filename': filename,
        'nomor': nomor_wa_clean,
        'pesan': pesan
    })

if __name__ == '__main__':
    debug_mode = os.getenv('FLASK_ENV', 'development') == 'development'
    port = int(os.getenv('PORT', 5000))
    app.run(debug=debug_mode, host='0.0.0.0', port=port)
