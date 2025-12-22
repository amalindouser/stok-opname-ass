"""
Script untuk import master barang dari TB_BARANG.xlsx ke SQLite
"""
import sqlite3
import os
from openpyxl import load_workbook
from datetime import datetime

DB_FILE = "data/stok_opname.db"
EXCEL_FILE = "data/TB_BARANG.xlsx"

if not os.path.exists(EXCEL_FILE):
    print(f"❌ File {EXCEL_FILE} tidak ditemukan")
    exit(1)

# Load Excel
print("📖 Loading Excel file...")
wb = load_workbook(EXCEL_FILE)
ws = wb.active

# Cari header columns
headers = {}
for col_idx, cell in enumerate(ws[1], start=1):
    if cell.value:
        headers[cell.value.strip().upper()] = col_idx

# Tentukan kolom yang diperlukan
kode_col = headers.get('KODE') or headers.get('KODE_BARANG')
nama_col = headers.get('NAMA') or headers.get('NAMA_BARANG')
kategori_col = headers.get('KATEGORI')
departemen_col = headers.get('DEPARTEMEN')
stok_col = headers.get('STOK')
satuan_col = headers.get('SATUAN')
hjual_col = headers.get('HJUAL')
hpp_col = headers.get('HPP')
margin_col = headers.get('MARGIN')
expdate_col = headers.get('EXPDATE')
status_promo_col = headers.get('STATUS_PROMO')

if not kode_col or not nama_col:
    print("❌ Error: Kolom KODE atau NAMA tidak ditemukan")
    print(f"Available columns: {list(headers.keys())}")
    exit(1)

# Connect ke database
conn = sqlite3.connect(DB_FILE)
c = conn.cursor()

# Create table jika belum ada
c.execute('''CREATE TABLE IF NOT EXISTS master_barang (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    kode TEXT UNIQUE NOT NULL,
    nama TEXT NOT NULL,
    kategori TEXT,
    departemen TEXT,
    stok INTEGER DEFAULT 0,
    satuan TEXT DEFAULT 'Unit',
    hjual REAL,
    hpp REAL,
    margin REAL,
    expdate TEXT,
    status_promo TEXT,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
)''')

# Clear existing data
c.execute("DELETE FROM master_barang")

# Baca data dari Excel dan insert ke database
data = []
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    try:
        kode = row[kode_col - 1] if kode_col <= len(row) and row[kode_col - 1] else None
        nama = row[nama_col - 1] if nama_col <= len(row) and row[nama_col - 1] else None
        
        if not kode or not nama:
            continue
        
        # Normalize kode
        kode = str(int(kode)) if isinstance(kode, (int, float)) else str(kode).strip()
        nama = str(nama).strip()
        
        # Ambil atribut lainnya
        kategori = row[kategori_col - 1] if kategori_col and kategori_col <= len(row) else None
        departemen = row[departemen_col - 1] if departemen_col and departemen_col <= len(row) else None
        stok = row[stok_col - 1] if stok_col and stok_col <= len(row) else 0
        satuan = row[satuan_col - 1] if satuan_col and satuan_col <= len(row) else 'Unit'
        hjual = row[hjual_col - 1] if hjual_col and hjual_col <= len(row) else None
        hpp = row[hpp_col - 1] if hpp_col and hpp_col <= len(row) else None
        margin = row[margin_col - 1] if margin_col and margin_col <= len(row) else None
        expdate = row[expdate_col - 1] if expdate_col and expdate_col <= len(row) else None
        status_promo = row[status_promo_col - 1] if status_promo_col and status_promo_col <= len(row) else None
        
        data.append((kode, nama, kategori, departemen, stok, satuan, hjual, hpp, margin, expdate, status_promo))
    
    except Exception as e:
        print(f"⚠️ Error di baris {row_idx}: {str(e)}")
        continue

# Insert ke database
if data:
    try:
        # Gunakan INSERT OR REPLACE untuk handle duplikat
        c.executemany('''
            INSERT OR REPLACE INTO master_barang (kode, nama, kategori, departemen, stok, satuan, hjual, hpp, margin, expdate, status_promo)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', data)
        conn.commit()
        print(f"✅ Import berhasil: {len(data)} item")
    except Exception as e:
        print(f"❌ Error saat insert: {str(e)}")
        conn.rollback()
else:
    print("⚠️ Tidak ada data untuk diimport")

# Verify
c.execute("SELECT COUNT(*) FROM master_barang")
count = c.fetchone()[0]
print(f"📊 Total item di database: {count}")

# Show sample
if count > 0:
    c.execute("SELECT kode, nama, kategori, stok FROM master_barang LIMIT 5")
    print("\n📋 Sample data:")
    for row in c.fetchall():
        print(f"  {row[0]} | {row[1]} | {row[2]} | Stok: {row[3]}")

conn.close()
